import io
import json
import uuid
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import get_current_user
from app.envelope import ok
from app.models import (
    ComponentConfig,
    PayrollRun,
    SalaryRegister,
    SalaryRegisterRow,
    TenantRulePreference,
    User,
)
from app.schemas.payroll import UploadParseResponse, ValidateRequest
from app.services.payroll_parse import (
    dataframe_to_employees,
    parse_payroll_file,
    validate_required_columns,
)
from app.services.validation import (
    _component_key_map,
    apply_suppressed_rules,
    split_row_amounts,
    validate_employees,
)

router = APIRouter()


def _to_first_of_month(d: date | None) -> date | None:
    if d is None:
        return None
    return d.replace(day=1)


def _suppressed_rule_ids(db: Session, user_id: uuid.UUID) -> set[str]:
    rows = (
        db.query(TenantRulePreference.rule_id)
        .filter(
            TenantRulePreference.user_id == user_id,
            TenantRulePreference.suppressed.is_(True),
        )
        .all()
    )
    return {r[0] for r in rows}


def _persist_salary_register(
    db: Session,
    user: User,
    period_month: date,
    filename: str | None,
    employees: list[dict],
    comps: list[ComponentConfig],
) -> None:
    comp_by_key = _component_key_map(comps)

    existing = (
        db.query(SalaryRegister)
        .filter(SalaryRegister.user_id == user.id, SalaryRegister.period_month == period_month)
        .first()
    )
    if existing:
        db.query(SalaryRegisterRow).filter(SalaryRegisterRow.register_id == existing.id).delete()
        existing.filename = filename
        existing.employee_count = len(employees)
        register = existing
    else:
        register = SalaryRegister(
            user_id=user.id,
            period_month=period_month,
            filename=filename,
            employee_count=len(employees),
        )
        db.add(register)
        db.flush()

    for row in employees:
        eid = (
            row.get("employee_id")
            or row.get("emp_id")
            or row.get("employee_code")
        )
        if eid is None:
            continue
        if isinstance(eid, float) and eid == int(eid):
            eid = str(int(eid))
        eid = str(eid).strip()
        if not eid:
            continue

        ename = row.get("employee_name") or row.get("name")
        if isinstance(ename, float):
            ename = None

        regular, arrear_by_base, inc_arrear_total = split_row_amounts(row, comp_by_key)
        components_json = {k: float(v) for k, v in regular.items()}
        arrears_json = {k: float(v) for k, v in arrear_by_base.items()}

        paid_days_raw = row.get("paid_days")
        lop_days_raw = row.get("lop_days") or row.get("lop")
        try:
            paid_days = Decimal(str(paid_days_raw)) if paid_days_raw not in (None, "") else None
        except Exception:
            paid_days = None
        try:
            lop_days = Decimal(str(lop_days_raw)) if lop_days_raw not in (None, "") else None
        except Exception:
            lop_days = None

        db.add(
            SalaryRegisterRow(
                register_id=register.id,
                user_id=user.id,
                period_month=period_month,
                employee_id=eid,
                employee_name=ename if isinstance(ename, str) else None,
                paid_days=paid_days,
                lop_days=lop_days,
                components=components_json,
                arrears=arrears_json,
                increment_arrear_total=inc_arrear_total,
            )
        )

    db.commit()


def _payload_after_validation(rows: list, findings_summary: dict) -> dict:
    all_findings: list = []
    for r in rows:
        all_findings.extend(r.get("findings", []))
    risk_list = [
        {
            "employee_id": r["employee_id"],
            "employee_name": r.get("employee_name"),
            "risk_score": r["risk_score"],
            "risk_level": r["risk_level"],
            "score_breakdown": r.get("score_breakdown", {}),
        }
        for r in rows
    ]
    return {
        "results": rows,
        "findings": all_findings,
        "findings_summary": findings_summary,
        "risk_scores": risk_list,
    }


@router.post("/upload")
async def upload_payroll(
    file: UploadFile = File(...),
    meta: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    try:
        payload = json.loads(meta)
        run_type = payload.get("run_type", "regular")
        eff_from = payload.get("effective_month_from")
        eff_to = payload.get("effective_month_to")
        period_month = payload.get("period_month")
        strict = payload.get("strict_header_check", True)
        eff_from_d = date.fromisoformat(eff_from) if eff_from else None
        eff_to_d = date.fromisoformat(eff_to) if eff_to else None
        period_month_d = date.fromisoformat(period_month) if period_month else None
    except (ValueError, json.JSONDecodeError):
        raise HTTPException(status_code=400, detail="Invalid meta JSON")

    raw = await file.read()
    try:
        df = parse_payroll_file(raw, file.filename or "upload.csv")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    columns, employees = dataframe_to_employees(df)
    comps = db.query(ComponentConfig).filter(ComponentConfig.user_id == user.id).all()
    comp_names = {c.component_name for c in comps}
    missing, warnings = validate_required_columns(columns, comp_names, strict=strict)

    preview = employees[:5]

    run = PayrollRun(
        user_id=user.id,
        run_type=run_type,
        effective_month_from=eff_from_d,
        effective_month_to=eff_to_d,
        filename=file.filename,
        employee_count=len(employees),
    )
    db.add(run)
    db.commit()

    persist_period = _to_first_of_month(period_month_d or eff_to_d)
    if persist_period and comps and not missing:
        _persist_salary_register(db, user, persist_period, file.filename, employees, comps)

    out = UploadParseResponse(
        columns=columns,
        preview=preview,
        employees=employees,
        missing_required=missing,
        warnings=warnings,
    )
    return ok(out.model_dump())


@router.post("/validate")
def validate_payroll(
    body: ValidateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    comps = db.query(ComponentConfig).filter(ComponentConfig.user_id == user.id).all()
    if not comps:
        raise HTTPException(status_code=400, detail="Configure salary components before validation.")
    period_month = _to_first_of_month(body.period_month or body.effective_month_to)
    rows, findings_summary = validate_employees(
        db,
        user,
        comps,
        body.employees,
        body.run_type,
        body.effective_month_from,
        body.effective_month_to,
        body.as_of_date,
        period_month=period_month,
    )
    suppressed = _suppressed_rule_ids(db, user.id)
    findings_summary = apply_suppressed_rules(rows, suppressed)
    return ok(_payload_after_validation(rows, findings_summary))


@router.get("/runs")
def list_payroll_runs(
    limit: int = 20,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    runs = (
        db.query(PayrollRun)
        .filter(PayrollRun.user_id == user.id)
        .order_by(PayrollRun.created_at.desc())
        .limit(limit)
        .all()
    )
    data = [
        {
            "id": str(r.id),
            "run_type": r.run_type,
            "filename": r.filename,
            "employee_count": r.employee_count,
            "effective_month_from": r.effective_month_from.isoformat() if r.effective_month_from else None,
            "effective_month_to": r.effective_month_to.isoformat() if r.effective_month_to else None,
            "created_at": r.created_at.isoformat(),
        }
        for r in runs
    ]
    return ok(data)


@router.get("/registers")
def list_salary_registers(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    regs = (
        db.query(SalaryRegister)
        .filter(SalaryRegister.user_id == user.id)
        .order_by(SalaryRegister.period_month.desc())
        .all()
    )
    data = [
        {
            "id": str(r.id),
            "period_month": r.period_month.isoformat(),
            "filename": r.filename,
            "employee_count": r.employee_count,
            "created_at": r.created_at.isoformat(),
        }
        for r in regs
    ]
    return ok(data)


@router.get("/registers/{register_id}")
def get_salary_register(
    register_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    try:
        rid = uuid.UUID(register_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Register not found")
    reg = (
        db.query(SalaryRegister).filter(SalaryRegister.id == rid, SalaryRegister.user_id == user.id).first()
    )
    if not reg:
        raise HTTPException(status_code=404, detail="Register not found")
    rows = (
        db.query(SalaryRegisterRow)
        .filter(SalaryRegisterRow.register_id == reg.id)
        .order_by(SalaryRegisterRow.employee_id)
        .all()
    )
    payload = {
        "id": str(reg.id),
        "period_month": reg.period_month.isoformat(),
        "filename": reg.filename,
        "employee_count": reg.employee_count,
        "created_at": reg.created_at.isoformat(),
        "rows": [
            {
                "employee_id": r.employee_id,
                "employee_name": r.employee_name,
                "paid_days": float(r.paid_days) if r.paid_days is not None else None,
                "lop_days": float(r.lop_days) if r.lop_days is not None else None,
                "components": r.components,
                "arrears": r.arrears,
                "increment_arrear_total": float(r.increment_arrear_total) if r.increment_arrear_total else 0,
            }
            for r in rows
        ],
    }
    return ok(payload)


@router.post("/validate/export-excel")
def export_findings_excel(
    body: ValidateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Run validation and return findings as an Excel workbook (binary stream, not JSON envelope)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    comps = db.query(ComponentConfig).filter(ComponentConfig.user_id == user.id).all()
    if not comps:
        raise HTTPException(status_code=400, detail="Configure salary components before export.")
    period_month = _to_first_of_month(body.period_month or body.effective_month_to)
    rows, findings_summary = validate_employees(
        db,
        user,
        comps,
        body.employees,
        body.run_type,
        body.effective_month_from,
        body.effective_month_to,
        body.as_of_date,
        period_month=period_month,
    )
    suppressed = _suppressed_rule_ids(db, user.id)
    findings_summary = apply_suppressed_rules(rows, suppressed)

    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1E293B")

    ws_sum.append(["Metric", "Value"])
    for cell in ws_sum[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    ws_sum.append(["Total Employees", len(rows)])
    ws_sum.append(["Total Findings", findings_summary.get("total_findings", 0)])
    ws_sum.append(["Critical", findings_summary.get("critical", 0)])
    ws_sum.append(["Warning", findings_summary.get("warning", 0)])
    ws_sum.append(["Info", findings_summary.get("info", 0)])
    ws_sum.append(["Passed", findings_summary.get("pass", 0)])
    ws_sum.append(["Total Financial Impact (₹)", findings_summary.get("total_financial_impact", 0)])
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    ws_f = wb.create_sheet("Findings")
    f_headers = [
        "Employee ID",
        "Employee Name",
        "Rule ID",
        "Rule Name",
        "Component",
        "Expected",
        "Actual",
        "Difference",
        "Severity",
        "Status",
        "Reason",
        "Suggested Fix",
        "Financial Impact (₹)",
    ]
    ws_f.append(f_headers)
    for cell in ws_f[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    severity_fills = {
        "CRITICAL": PatternFill("solid", fgColor="FEE2E2"),
        "WARNING": PatternFill("solid", fgColor="FEF9C3"),
        "INFO": PatternFill("solid", fgColor="EFF6FF"),
    }
    for emp in rows:
        for f in emp.get("findings", []):
            row_data = [
                f.get("employee_id", ""),
                f.get("employee_name", ""),
                f.get("rule_id", ""),
                f.get("rule_name", ""),
                f.get("component", ""),
                f.get("expected_value", ""),
                f.get("actual_value", ""),
                f.get("difference", ""),
                f.get("severity", ""),
                f.get("status", ""),
                f.get("reason", ""),
                f.get("suggested_fix", ""),
                f.get("financial_impact", 0),
            ]
            ws_f.append(row_data)
            if f.get("status") == "FAIL":
                sev = f.get("severity", "")
                fill = severity_fills.get(sev)
                if fill:
                    for cell in ws_f[ws_f.max_row]:
                        cell.fill = fill

    for col in ws_f.columns:
        ws_f.column_dimensions[col[0].column_letter].width = 22
    ws_f.column_dimensions["K"].width = 60
    ws_f.column_dimensions["L"].width = 60

    ws_r = wb.create_sheet("Risk Scores")
    ws_r.append(["Employee ID", "Employee Name", "Risk Score", "Risk Level"])
    for cell in ws_r[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    level_fills = {
        "HIGH": PatternFill("solid", fgColor="FEE2E2"),
        "MEDIUM": PatternFill("solid", fgColor="FEF9C3"),
        "LOW": PatternFill("solid", fgColor="F0FDF4"),
    }
    for emp in rows:
        ws_r.append(
            [
                emp.get("employee_id", ""),
                emp.get("employee_name", ""),
                emp.get("risk_score", 0),
                emp.get("risk_level", "LOW"),
            ]
        )
        lvl = emp.get("risk_level", "LOW")
        if lvl in level_fills:
            for cell in ws_r[ws_r.max_row]:
                cell.fill = level_fills[lvl]

    for col in ["A", "B", "C", "D"]:
        ws_r.column_dimensions[col].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payroll-audit-report.xlsx"},
    )


@router.get("/dashboard-stats")
def dashboard_stats(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    n_comp = db.query(ComponentConfig).filter(ComponentConfig.user_id == user.id).count()
    last_run = (
        db.query(PayrollRun)
        .filter(PayrollRun.user_id == user.id)
        .order_by(PayrollRun.created_at.desc())
        .first()
    )
    last_register = (
        db.query(SalaryRegister)
        .filter(SalaryRegister.user_id == user.id)
        .order_by(SalaryRegister.period_month.desc())
        .first()
    )
    payload = {
        "components_configured": n_comp,
        "last_run_employee_count": last_run.employee_count if last_run else 0,
        "last_run_at": last_run.created_at.isoformat() if last_run else None,
        "last_register_period": last_register.period_month.isoformat() if last_register else None,
        "message": "Upload and validate payroll to populate PF/ESIC stats.",
    }
    return ok(payload)
